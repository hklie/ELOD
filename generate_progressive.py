#!/usr/bin/env python3
"""
Generate progressive ELOD Excel file showing delta per player per tournament.
"""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Import from elod module
from elod import (
    Elod, parse_manifest, parse_aliases, apply_aliases,
    TournamentData, DEFAULT_ELO
)
from player import Player


def load_deceased_players(file_path: str) -> Set[str]:
    """Load list of deceased players to exclude from rankings."""
    deceased = set()
    path = Path(file_path)
    if not path.exists():
        return deceased

    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#'):
                deceased.add(line)

    return deceased


def load_country_data(file_path: str) -> Dict[str, str]:
    """Load country of origin for each player."""
    countries = {}
    path = Path(file_path)
    if not path.exists():
        return countries

    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                canonical, country = line.split('=', 1)
                countries[canonical.strip()] = country.strip()

    return countries


def load_display_names(file_path: str) -> Dict[str, str]:
    """Load custom display names for players with compound surnames."""
    display_names = {}
    path = Path(file_path)
    if not path.exists():
        return display_names

    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                canonical, display = line.split('=', 1)
                display_names[canonical.strip()] = display.strip()

    return display_names


def name_to_lastname_firstname(name: str, display_names: Optional[Dict[str, str]] = None) -> str:
    """
    Convert CamelCase name to "LASTNAME, Firstname" format.
    Uses display_names dictionary for special cases.
    """
    # Check for custom display name first
    if display_names and name in display_names:
        return display_names[name]

    # Find the split point between firstname and lastname
    parts = re.findall(r'[A-ZÁÉÍÓÚÑÜ][a-záéíóúñü]+', name)

    if len(parts) >= 2:
        if len(parts) >= 3:
            firstname = ''.join(parts[:-1])
            lastname = parts[-1]
        else:
            firstname = parts[0]
            lastname = parts[1]
        return f"{lastname.upper()}, {firstname}"

    return name.upper()


class ProgressiveElod:
    """Track ELOD deltas per player per tournament."""

    def __init__(self):
        self.players: Dict[str, Player] = {}
        self.tournament_deltas: Dict[str, Dict[str, float]] = {}
        self.tournament_order: List[str] = []
        self.tournament_participants: Dict[str, int] = {}  # Participant count per partida
        self.initial_elos: Dict[str, float] = {}
        self.deceased_players: Set[str] = set()
        self.display_names: Dict[str, str] = {}
        self.country_data: Dict[str, str] = {}

    def process_tournaments(
        self,
        manifest_path: str,
        alias_path: Optional[str] = None,
        base_path: Optional[str] = None,
        deceased_path: Optional[str] = None,
        display_names_path: Optional[str] = None,
        country_path: Optional[str] = None
    ):
        """Process all tournaments and track deltas."""

        # Load aliases
        aliases = {}
        if alias_path:
            aliases = parse_aliases(alias_path)
            print(f"Loaded {len(aliases)} aliases")

        # Load deceased players list
        if deceased_path:
            self.deceased_players = load_deceased_players(deceased_path)
            print(f"Loaded {len(self.deceased_players)} deceased players to exclude")

        # Load display names
        if display_names_path:
            self.display_names = load_display_names(display_names_path)
            print(f"Loaded {len(self.display_names)} custom display names")

        # Load country data
        if country_path:
            self.country_data = load_country_data(country_path)
            print(f"Loaded {len(self.country_data)} player countries")

        # Load tournament list
        tournaments = parse_manifest(manifest_path, base_path)
        print(f"Loaded {len(tournaments)} tournament files")

        # Group tournaments by display name
        tournament_groups: Dict[str, List[Tuple[str, str]]] = OrderedDict()
        for filepath, display_name in tournaments:
            if display_name not in tournament_groups:
                tournament_groups[display_name] = []
            tournament_groups[display_name].append((filepath, display_name))

        print(f"Found {len(tournament_groups)} unique tournaments")

        # Don't pass base_path to Elod - parse_manifest already resolved paths
        elod = Elod()

        # Initialize all players first
        for filepath, display_name in tournaments:
            elod.read_players_from_tournament(filepath, DEFAULT_ELO, aliases)

        # Store initial ELOs
        self.initial_elos = {name: player.elo for name, player in elod.players.items()}

        # Create working copy (like original elod.py)
        players_updated = {name: player.copy() for name, player in elod.players.items()}

        # Process each tournament, tracking each partida separately
        for tournament_name, files in tournament_groups.items():
            num_files = len(files)
            is_mundial = 'mundial' in tournament_name.lower()
            print(f"Processing: {tournament_name} ({num_files} partida{'s' if num_files > 1 else ''})")

            # Process each file in the tournament group
            for file_idx, (filepath, _) in enumerate(files, 1):
                # Store pre-partida ELO
                pre_partida_elo = {name: player.elo for name, player in elod.players.items()}

                tournament_data = elod.load_tournament(filepath, tournament_name)

                # Apply aliases
                tournament_data.players = [
                    apply_aliases(name, aliases) for name in tournament_data.players
                ]
                tournament_data.stopped_players = [
                    apply_aliases(name, aliases) for name in tournament_data.stopped_players
                ]

                # Track which players are in each category for this partida
                # If a player appears in both (data issue), prioritize full players (Duplicada)
                full_players_set = set(tournament_data.players)
                stopped_players_set = set(tournament_data.stopped_players) - full_players_set

                # Add any new players
                for player_name in tournament_data.all_players:
                    if player_name not in elod.players:
                        elod.players[player_name] = Player(
                            elo=DEFAULT_ELO, initial_elo=DEFAULT_ELO,
                            games=0, tourneys=0, last_tourney="-"
                        )
                        players_updated[player_name] = elod.players[player_name].copy()
                        pre_partida_elo[player_name] = DEFAULT_ELO
                        self.initial_elos[player_name] = DEFAULT_ELO

                # Set initial_elo for participating players
                for player_name in tournament_data.all_players:
                    if player_name in players_updated:
                        players_updated[player_name].initial_elo = players_updated[player_name].elo

                # Process tournament
                elod.process_tournament(tournament_data, players_updated)

                # Sync updates back after EACH FILE (like original elod.py)
                for name in elod.players:
                    if name in players_updated:
                        elod.players[name].elo = players_updated[name].elo
                        elod.players[name].games = players_updated[name].games
                        elod.players[name].tourneys = players_updated[name].tourneys
                        elod.players[name].last_tourney = players_updated[name].last_tourney

                # For Mundial tournaments P1 and P2, separate Duplicada and Copa FILE
                # P3 is only played by Duplicada players (Copa FILE stops at R15 of P2)
                is_p1_or_p2 = file_idx <= 2
                if is_mundial and stopped_players_set and is_p1_or_p2:
                    # Generate column names for both modalities
                    dup_name = f"{tournament_name} Dup P{file_idx}"
                    copa_name = f"{tournament_name} Copa P{file_idx}"

                    # Calculate deltas for Duplicada (full players)
                    self.tournament_deltas[dup_name] = {}
                    for name, player in elod.players.items():
                        if name in full_players_set:
                            delta = player.elo - pre_partida_elo.get(name, DEFAULT_ELO)
                            if abs(delta) > 0.001:
                                self.tournament_deltas[dup_name][name] = delta

                    # Calculate deltas for Copa FILE (stopped players)
                    self.tournament_deltas[copa_name] = {}
                    for name, player in elod.players.items():
                        if name in stopped_players_set:
                            delta = player.elo - pre_partida_elo.get(name, DEFAULT_ELO)
                            if abs(delta) > 0.001:
                                self.tournament_deltas[copa_name][name] = delta

                    # Track participant counts
                    self.tournament_participants[dup_name] = len(full_players_set)
                    self.tournament_participants[copa_name] = len(stopped_players_set)

                    self.tournament_order.append(dup_name)
                    self.tournament_order.append(copa_name)
                elif is_mundial and file_idx == 3:
                    # Mundial P3 - only Duplicada players (Copa FILE doesn't play P3)
                    partida_name = f"{tournament_name} Dup P3"

                    self.tournament_deltas[partida_name] = {}
                    for name, player in elod.players.items():
                        delta = player.elo - pre_partida_elo.get(name, DEFAULT_ELO)
                        if abs(delta) > 0.001:
                            self.tournament_deltas[partida_name][name] = delta

                    # Track participant count (only full/Duplicada players participate in P3)
                    self.tournament_participants[partida_name] = len(full_players_set)

                    self.tournament_order.append(partida_name)
                else:
                    # Non-Mundial tournament - single column
                    if num_files > 1:
                        partida_name = f"{tournament_name} P{file_idx}"
                    else:
                        partida_name = tournament_name

                    # Calculate deltas for this partida
                    self.tournament_deltas[partida_name] = {}
                    for name, player in elod.players.items():
                        delta = player.elo - pre_partida_elo.get(name, DEFAULT_ELO)
                        if abs(delta) > 0.001:
                            self.tournament_deltas[partida_name][name] = delta

                    # Track participant count
                    self.tournament_participants[partida_name] = len(tournament_data.all_players)

                    self.tournament_order.append(partida_name)

        # Store final player data
        self.players = elod.players

    def generate_excel(self, output_path: str):
        """Generate Excel file with progressive deltas."""

        wb = Workbook()
        ws = wb.active
        ws.title = "ELOD Progresivos"

        # Spanish month names for dynamic date
        MESES = {
            1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
            5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
            9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
        }
        now = datetime.now()
        fecha = f"{now.day} de {MESES[now.month]} de {now.year}"

        # Styles
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # --- Title rows (rows 1-3) ---
        tournament_columns = list(reversed(self.tournament_order))
        total_cols = 8 + len(tournament_columns)

        # Row 1: Institution name
        title_cell = ws.cell(row=1, column=1, value="Federación Internacional de Léxico en Español (FILE)")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_align
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        # Row 2: Subtitle left + date right
        subtitle_cell = ws.cell(row=2, column=1, value="Elo Duplicada - Ranking Internacional")
        subtitle_cell.font = Font(bold=True, size=12)
        subtitle_cell.alignment = center_align
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

        date_cell = ws.cell(row=2, column=5, value=f"Actualizado al {fecha}")
        date_cell.font = Font(bold=True, size=12)
        date_cell.alignment = center_align
        ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=8)

        # Row 3: Empty separator

        # --- Column headers (row 4) ---
        # Column order: A=Deltas acum, B=Pos, C=JUGADOR, D=País, E=ELOD Actual,
        #               F=Último Torneo, G=N.Oponentes, H=N.Partidas, I+=tournaments
        headers = ["Deltas acumulados", "Pos.", "JUGADOR", "País", "ELOD Actual",
                    "Último Torneo", "N.Oponentes", "N.Partidas"]
        headers.extend(tournament_columns)
        tournament_start_col = 9

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Vertical text for tournament column headers
        vertical_align = Alignment(horizontal='center', vertical='center', textRotation=90)
        for col in range(tournament_start_col, len(headers) + 1):
            ws.cell(row=4, column=col).alignment = vertical_align

        # --- Participant counts row (row 5) ---
        participant_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        for col in range(1, 3):  # Columns A-B empty
            ws.cell(row=5, column=col, value="").fill = participant_fill
        cell = ws.cell(row=5, column=3, value="Participantes")
        cell.font = Font(bold=True, italic=True)
        cell.fill = participant_fill
        for col in range(4, 9):  # Columns D-H empty
            ws.cell(row=5, column=col, value="").fill = participant_fill

        for col, tournament in enumerate(tournament_columns, tournament_start_col):
            count = self.tournament_participants.get(tournament, 0)
            cell = ws.cell(row=5, column=col, value=count)
            cell.alignment = center_align
            cell.fill = participant_fill
            cell.font = Font(italic=True)

        # --- Player data (row 6+) ---
        sorted_players = sorted(
            [(name, player) for name, player in self.players.items()
             if name not in self.deceased_players],
            key=lambda x: x[1].elo,
            reverse=True
        )

        if self.deceased_players:
            excluded = len(self.players) - len(sorted_players)
            print(f"Excluded {excluded} deceased players from rankings")

        for row, (name, player) in enumerate(sorted_players, 6):
            # A: Cumulative delta
            cumulative_delta = 0
            for tournament in self.tournament_order:
                delta = self.tournament_deltas.get(tournament, {}).get(name)
                if delta is not None:
                    cumulative_delta += round(delta)
            ws.cell(row=row, column=1, value=cumulative_delta).alignment = center_align

            # B: Position (row-5 because rows 1-3=title, 4=header, 5=participants)
            ws.cell(row=row, column=2, value=f"{row-5}°").alignment = center_align

            # C: JUGADOR
            formatted_name = name_to_lastname_firstname(name, self.display_names)
            ws.cell(row=row, column=3, value=formatted_name)

            # D: País
            country = self.country_data.get(name, "")
            ws.cell(row=row, column=4, value=country)

            # E: ELOD Actual
            ws.cell(row=row, column=5, value=round(player.elo)).alignment = center_align

            # F: Último Torneo
            ws.cell(row=row, column=6, value=player.last_tourney).alignment = center_align

            # G: N.Oponentes
            ws.cell(row=row, column=7, value=player.games).alignment = center_align

            # H: N.Partidas
            ws.cell(row=row, column=8, value=player.tourneys).alignment = center_align

            # I+: Tournament deltas
            for col, tournament in enumerate(tournament_columns, tournament_start_col):
                delta = self.tournament_deltas.get(tournament, {}).get(name)
                if delta is not None:
                    cell = ws.cell(row=row, column=col, value=round(delta))
                    cell.alignment = center_align
                    if delta < 0:
                        cell.font = Font(color="FF0000")
                    elif delta > 0:
                        cell.font = Font(color="008000")

        # Column widths
        ws.column_dimensions['A'].width = 12  # Deltas acumulados
        ws.column_dimensions['B'].width = 6   # Pos.
        ws.column_dimensions['C'].width = 25  # JUGADOR
        ws.column_dimensions['D'].width = 18  # País
        ws.column_dimensions['E'].width = 10  # ELOD Actual
        ws.column_dimensions['F'].width = 28  # Último Torneo
        ws.column_dimensions['G'].width = 12  # N.Oponentes
        ws.column_dimensions['H'].width = 10  # N.Partidas
        for col in range(tournament_start_col, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 5

        # Freeze panes (freeze first 3 columns and first 5 rows)
        ws.freeze_panes = 'D6'

        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Generate progressive ELOD Excel file')
    parser.add_argument('--manifest', '-m', required=True, help='Tournament manifest file')
    parser.add_argument('--aliases', '-a', help='Player aliases file')
    parser.add_argument('--output', '-o', default='./output/elod_progresivos.xlsx', help='Output Excel file')
    parser.add_argument('--base-path', '-b', help='Base path for tournament files')
    parser.add_argument('--deceased', '-d', help='File with deceased players to exclude')
    parser.add_argument('--display-names', '-n', help='File with custom display names')
    parser.add_argument('--country', '-c', help='File with player country data')

    args = parser.parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    prog = ProgressiveElod()
    prog.process_tournaments(
        manifest_path=args.manifest,
        alias_path=args.aliases,
        base_path=args.base_path,
        deceased_path=args.deceased,
        display_names_path=args.display_names,
        country_path=args.country
    )
    prog.generate_excel(str(output_path))


if __name__ == '__main__':
    main()
